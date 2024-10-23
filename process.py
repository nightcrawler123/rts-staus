import pandas as pd
import os
from openpyxl import load_workbook

# Load the Excel file
excel_file = "NA Trend Report.xlsx"
wb = load_workbook(excel_file)

# Get all sheet names
sheet_names = wb.sheetnames

# Delete rows with the oldest date from each sheet
for sheet_name in sheet_names:
    sheet = wb[sheet_name]
    # Skip if the sheet is empty or only contains headers
    if sheet.max_row < 2:
        continue
    
    # Collect dates from the first column
    dates = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]
    # Filter only valid dates and find the oldest one
    dates = [d for d in dates if isinstance(d, (pd.Timestamp, pd.datetime))]
    
    if dates:
        oldest_date = min(dates)
        
        # Delete all rows with the oldest date in the first column
        rows_to_delete = []
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == oldest_date:
                rows_to_delete.append(row)
        
        for row in reversed(rows_to_delete):  # Reverse to avoid shifting
            sheet.delete_rows(row)

# Save the workbook after deleting the rows
wb.save(excel_file)

# CSV to tab mapping
csv_to_tab = {
    'QDS-above-70-crossed-40d': 'QDS above 70 G40',
    'QDS-0-69-crossed-40d': 'QDS below 70 G40',
    'QDS-0-69-less-40d': 'QDS below 70 L40',
    'QDS-above-70-less-40d': 'QDS above 70 L40'
}

# Find all processed CSV files in the current directory
csv_files = [f for f in os.listdir() if f.startswith('processed') and f.endswith('.csv')]

for csv_file in csv_files:
    for pattern, tab_name in csv_to_tab.items():
        if pattern in csv_file:
            # Load the CSV, skipping the header
            csv_data = pd.read_csv(csv_file, skiprows=1)
            
            # Load the relevant sheet from the Excel file
            sheet = wb[tab_name]
            
            # Append the CSV data to the sheet
            for row in csv_data.itertuples(index=False):
                sheet.append(list(row))
            
            print(f"Data from {csv_file} appended to sheet {tab_name}")
            break  # Stop after finding the matching tab

# Save the workbook after copying data
wb.save(excel_file)
