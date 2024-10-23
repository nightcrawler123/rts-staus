import sys
import site
import pandas as pd
import glob
import os
from tqdm import tqdm
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import gc

# Ensure user site-packages are included
def add_user_site_packages():
    try:
        user_site = site.getusersitepackages()
        if user_site not in sys.path:
            sys.path.append(user_site)
    except AttributeError:
        # Fallback for environments where getusersitepackages is not available
        user_base = site.getuserbase()
        user_site = os.path.join(user_base, 'lib', f'python{sys.version_info.major}.{sys.version_info.minor}', 'site-packages')
        if user_site not in sys.path:
            sys.path.append(user_site)

add_user_site_packages()

def delete_oldest_date_rows(sheet_df):
    """
    Delete rows from the DataFrame that have the oldest date in the first column.
    Converts all columns to strings to handle mixed-type columns.
    """
    if not sheet_df.empty:
        # Convert all columns to string to handle mixed types
        sheet_df = sheet_df.astype(str)
        
        # Convert the first column to datetime, coerce errors to NaT
        sheet_df.iloc[:, 0] = pd.to_datetime(sheet_df.iloc[:, 0], errors='coerce', dayfirst=True)
        # Drop rows where the first column is NaT
        sheet_df.dropna(subset=[sheet_df.columns[0]], inplace=True)
        if sheet_df.empty:
            return sheet_df  # Return empty DataFrame if no valid dates
        # Find the oldest date
        oldest_date = sheet_df.iloc[:, 0].min()
        # Remove rows with the oldest date
        sheet_df = sheet_df[sheet_df.iloc[:, 0] != oldest_date]
    return sheet_df

def map_csv_to_sheet(csv_filename):
    """
    Map CSV file name patterns to corresponding Excel sheet names.
    """
    if "QDS-above-70-crossed-40d" in csv_filename:
        return "QDS above 70 G40"
    elif "QDS-0-69-crossed-40d" in csv_filename:
        return "QDS below 70 G40"
    elif "QDS-0-69-less-40d" in csv_filename:
        return "QDS below 70 L40"
    elif "QDS-above-70-less-40d" in csv_filename:
        return "QDS above 70 L40"
    else:
        return None  # No matching sheet

def process_excel_sheet(sheet_name, input_workbook):
    """
    Process a single Excel sheet: delete oldest date rows.
    """
    ws = input_workbook[sheet_name]
    # Read the sheet data into a DataFrame
    data = ws.values
    try:
        columns = next(data)  # Assumes first row is header
    except StopIteration:
        # Empty sheet
        return pd.DataFrame()  # Return empty DataFrame
    
    df = pd.DataFrame(data, columns=columns)
    
    # Delete rows with the oldest date
    df = delete_oldest_date_rows(df)
    
    # Convert all columns to strings
    df = df.astype(str)
    
    # Return the processed DataFrame
    return df

def append_csv_data(df, csv_file):
    """
    Append data from a CSV file to the DataFrame.
    """
    try:
        # Read CSV in chunks to handle large files
        chunk_size = 10000  # Adjust based on memory and performance
        for chunk in pd.read_csv(csv_file, chunksize=chunk_size, dtype=str):
            # If DataFrame already has data, skip header row in CSV
            if not df.empty:
                chunk = chunk.iloc[1:]  # Skip header row
            if chunk.empty:
                continue  # Skip empty chunks
            df = pd.concat([df, chunk], ignore_index=True)
            # Force garbage collection after each chunk
            del chunk
            gc.collect()
    except pd.errors.EmptyDataError:
        # Handle empty CSV files gracefully
        pass
    return df

def main():
    # Display current working directory
    print(f"Current working directory: {os.getcwd()}")
    
    # Display current time at the start
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"Script started at: {current_time}")
    
    # Start time for total execution time
    start_time = time.time()
    
    # Define the input and output Excel file paths
    input_excel_file = 'NA Trend Report.xlsx'
    output_excel_file = 'NA Trend Report Updated.xlsx'
    
    # Verify that the input Excel file exists
    if not os.path.isfile(input_excel_file):
        print(f"Input Excel file '{input_excel_file}' not found in the current directory.")
        return
    
    # Find all CSV files in the current directory
    csv_files = glob.glob('*.csv')  # Changed from 'processed_*.csv' to '*.csv'
    
    # Debugging: Print the list of CSV files found
    print(f"CSV files found: {csv_files}")
    
    if not csv_files:
        print("No CSV files found. Exiting script.")
        return
    
    # Create a mapping from sheet names to their corresponding CSV files
    sheet_csv_mapping = {}
    for csv_file in csv_files:
        sheet_name = map_csv_to_sheet(os.path.basename(csv_file))
        if sheet_name:
            if sheet_name not in sheet_csv_mapping:
                sheet_csv_mapping[sheet_name] = []
            sheet_csv_mapping[sheet_name].append(csv_file)
    
    # Debugging: Print the mapping
    print("Sheet to CSV mapping:")
    for sheet, files in sheet_csv_mapping.items():
        print(f"  {sheet}: {files}")
    
    if not sheet_csv_mapping:
        print("No CSV files match the expected naming patterns. Exiting script.")
        return
    
    # Load the input workbook
    print("Loading the Excel workbook...")
    try:
        input_wb = load_workbook(input_excel_file, read_only=True, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return
    
    sheet_names = input_wb.sheetnames
    print(f"Sheets in workbook: {sheet_names}")
    
    # Initialize the output workbook using pandas ExcelWriter
    with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:
        # Iterate through each sheet and process
        for sheet_name in tqdm(sheet_names, desc="Processing Excel Sheets", unit="sheet"):
            # Process the sheet to delete oldest date rows
            df = process_excel_sheet(sheet_name, input_wb)
            
            # If there are corresponding CSV files, append their data
            if sheet_name in sheet_csv_mapping:
                for csv_file in sheet_csv_mapping[sheet_name]:
                    print(f"Appending data from '{csv_file}' to sheet '{sheet_name}'")
                    df = append_csv_data(df, csv_file)
            
            # Write the processed DataFrame to the new Excel workbook
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Force garbage collection after each sheet
            del df
            gc.collect()
    
    # Close the input workbook
    input_wb.close()
    
    # Calculate and display the total time taken
    end_time = time.time()
    total_time = end_time - start_time
    print(f"Processing completed in {total_time:.2f} seconds.")
    
    # Display end time
    end_time_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"Script ended at: {end_time_datetime}")

if __name__ == "__main__":
    main()
