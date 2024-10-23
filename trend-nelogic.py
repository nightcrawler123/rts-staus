import sys
import subprocess
import importlib
import os
import glob
import logging
from datetime import datetime
import shutil
from tqdm import tqdm
import gc  # For garbage collection

# List of required packages
required_packages = [
    'polars',
    'tqdm',
    'openpyxl',
    'pyarrow',
    'pandas'  # Needed for Excel writing compatibility
]

# Function to install missing packages
def install_packages(packages):
    for package in packages:
        try:
            importlib.import_module(package)
        except ImportError:
            print(f"Package '{package}' not found. Installing...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# Install missing packages
install_packages(required_packages)

# Now import the installed packages
import polars as pl
import pandas as pd

# Setup logging at the very beginning to capture all events
logging.basicConfig(
    filename='data_processing.log',
    filemode='w',  # Overwrite log file each run
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

from openpyxl import load_workbook

def mangle_duplicate_columns(columns):
    """
    Makes duplicate column names unique by appending _1, _2, etc.
    """
    seen = {}
    new_columns = []
    for col in columns:
        if col in seen:
            seen[col] += 1
            new_col = f"{col}_{seen[col]}"
        else:
            seen[col] = 0
            new_col = col
        new_columns.append(new_col)
    return new_columns

def map_csv_to_sheet(csv_filename):
    """
    Maps a CSV filename to the corresponding Excel sheet name based on predefined patterns.
    """
    basename = os.path.splitext(csv_filename)[0]
    mapping = {
        'QDS-above-70-crossed-40d': 'QDS above 70 G40',
        'QDS-0-69-crossed-40d': 'QDS below 70 G40',
        'QDS-0-69-less-40d': 'QDS below 70 L40',
        'QDS-above-70-less-40d': 'QDS above 70 L40'
    }
    return mapping.get(basename, None)

def read_excel_sheets(excel_path):
    """
    Reads all sheets from the Excel file into a dictionary of Polars DataFrames using openpyxl in read-only mode.
    Handles duplicate column names by mangling them to make unique.
    Handles missing values and special characters.
    """
    try:
        print(f"Opening Excel file '{excel_path}' with openpyxl...")
        logging.info(f"Opening Excel file '{excel_path}' with openpyxl...")
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        print("Excel file opened successfully.")
        sheet_names = wb.sheetnames
        print(f"Found sheets: {sheet_names}")
        logging.info(f"Found sheets: {sheet_names}")
        polars_dict = {}
        for sheet_name in tqdm(sheet_names, desc="Reading Excel Sheets"):
            print(f"Reading sheet: {sheet_name}")
            sheet_start_time = datetime.now()
            ws = wb[sheet_name]

            # Extract data using generator to minimize memory usage
            data = ws.values
            try:
                # Get the first line as columns header
                columns = next(data)
            except StopIteration:
                # Handle empty sheets
                print(f"Sheet '{sheet_name}' is empty. Skipping.")
                logging.warning(f"Sheet '{sheet_name}' is empty. Skipping.")
                polars_dict[sheet_name] = pl.DataFrame()
                continue

            # Mangle duplicate column names
            columns = mangle_duplicate_columns(columns)

            # Convert the rest of the data into a list of rows
            data_rows = list(data)

            # Create Polars DataFrame with explicit schema
            try:
                df = pl.DataFrame(data_rows, schema=columns)
            except Exception as e:
                logging.error(f"Error creating Polars DataFrame for sheet '{sheet_name}': {e}")
                print(f"Error creating DataFrame for sheet '{sheet_name}': {e}")
                polars_dict[sheet_name] = pl.DataFrame()
                continue

            # Handle Missing Values
            df = handle_missing_values(df)

            # Handle Special Characters
            df = handle_special_characters(df)

            # Handle Duplicate Rows
            df = handle_duplicate_rows(df, sheet_name)

            # Log missing values
            log_missing_values(df, sheet_name)

            # Assign to dictionary
            polars_dict[sheet_name] = df

            sheet_end_time = datetime.now()
            elapsed_time = sheet_end_time - sheet_start_time
            print(f"Finished reading sheet: {sheet_name} in {elapsed_time}")
            logging.info(f"Finished reading sheet: {sheet_name} in {elapsed_time}")

            # Free memory
            del ws
            del data_rows
            del data
            gc.collect()
        print("Completed reading all Excel sheets.")
        wb.close()
        return polars_dict
    except Exception as e:
        logging.error(f"Error reading Excel file '{excel_path}': {e}")
        sys.exit(f"Error reading Excel file '{excel_path}': {e}")

def handle_missing_values(df):
    """
    Handles missing values in the DataFrame.
    For numerical columns, fills missing values with the mean.
    For categorical/textual missing values, fills with 'Unknown'.
    """
    try:
        for col in df.columns:
            if df[col].dtype in [pl.Float64, pl.Int64, pl.UInt64]:
                # Fill numerical missing values with mean
                mean_val = df[col].mean()
                df = df.with_column(
                    pl.col(col).fill_null(mean_val)
                )
            elif df[col].dtype == pl.Date:
                # Fill missing dates with a default date, e.g., 1970-01-01
                df = df.with_column(
                    pl.col(col).fill_null(datetime(1970, 1, 1)).cast(pl.Date)
                )
            else:
                # Fill categorical/text missing values with 'Unknown'
                df = df.with_column(
                    pl.col(col).fill_null("Unknown")
                )
        return df
    except Exception as e:
        logging.error(f"Error handling missing values: {e}")
        print(f"Error handling missing values: {e}")
        return df

def handle_special_characters(df):
    """
    Ensures that special characters in string columns are properly handled.
    Strips unwanted characters and preserves necessary ones.
    """
    try:
        string_columns = [col for col in df.columns if df[col].dtype == pl.Utf8]
        for col in string_columns:
            # Example: Strip leading/trailing whitespace and replace problematic characters
            df = df.with_column(
                pl.col(col).str.strip().str.replace(r'[^\x00-\x7F]+', '', strict=False).alias(col)
            )
        return df
    except Exception as e:
        logging.error(f"Error handling special characters: {e}")
        print(f"Error handling special characters: {e}")
        return df

def handle_duplicate_rows(df, sheet_name):
    """
    Removes duplicate rows from the DataFrame.
    """
    try:
        initial_count = df.height
        df = df.unique()
        final_count = df.height
        duplicates_removed = initial_count - final_count
        if duplicates_removed > 0:
            logging.info(f"Sheet '{sheet_name}': Removed {duplicates_removed} duplicate rows.")
            print(f"Sheet '{sheet_name}': Removed {duplicates_removed} duplicate rows.")
        return df
    except Exception as e:
        logging.error(f"Error handling duplicate rows in sheet '{sheet_name}': {e}")
        print(f"Error handling duplicate rows in sheet '{sheet_name}': {e}")
        return df

def log_missing_values(df, sheet_name):
    """
    Logs the count of missing values per column in the DataFrame.
    """
    try:
        missing_counts = df.null_count().to_dict()
        for col, count in missing_counts.items():
            if count > 0:
                logging.info(f"Sheet '{sheet_name}': Column '{col}' has {count} missing values.")
    except Exception as e:
        logging.error(f"Error logging missing values for sheet '{sheet_name}': {e}")

def read_csv_file(csv_path):
    """
    Reads a CSV file into a Polars DataFrame with all columns as strings.
    Handles missing values and special characters.
    Logs any malformed rows or read errors.
    """
    try:
        print(f"Reading CSV file: {csv_path}")
        df = pl.read_csv(
            csv_path,
            try_parse_dates=False,      # Disable automatic date parsing
            ignore_errors=False,        # Do not ignore errors to capture bad rows
            low_memory=False,           # Disable low_memory to improve type inference
            n_threads=4,                # Utilize all available cores
            infer_schema_length=1000,   # Infer schema from first 1000 rows
            dtype={"Date": pl.Utf8},    # Ensure Date column is read as string
            encoding='utf-8',           # Handle special characters
            has_header=True
        )
        logging.info(f"Successfully read CSV file '{csv_path}' with {df.height} rows and {df.width} columns.")
        print(f"Successfully read CSV file: {csv_path} with {df.height} rows and {df.width} columns.")

        # Handle Missing Values
        df = handle_missing_values(df)

        # Handle Special Characters
        df = handle_special_characters(df)

        # Handle Duplicate Rows
        df = handle_duplicate_rows_csv(df, csv_path)

        # Log missing values
        log_missing_values_csv(df, csv_path)

        return df
    except pl.errors.ParseError as pe:
        logging.error(f"Parse error in file '{csv_path}': {pe}")
        print(f"Parse error in file '{csv_path}': {pe}")
    except Exception as e:
        logging.error(f"Unexpected error reading CSV file '{csv_path}': {e}")
        print(f"Unexpected error reading CSV file '{csv_path}': {e}")
    return None

def handle_duplicate_rows_csv(df, csv_path):
    """
    Removes duplicate rows from the CSV DataFrame.
    """
    try:
        initial_count = df.height
        df = df.unique()
        final_count = df.height
        duplicates_removed = initial_count - final_count
        if duplicates_removed > 0:
            logging.info(f"CSV '{csv_path}': Removed {duplicates_removed} duplicate rows.")
            print(f"CSV '{csv_path}': Removed {duplicates_removed} duplicate rows.")
        return df
    except Exception as e:
        logging.error(f"Error handling duplicate rows in CSV '{csv_path}': {e}")
        print(f"Error handling duplicate rows in CSV '{csv_path}': {e}")
        return df

def log_missing_values_csv(df, csv_path):
    """
    Logs the count of missing values per column in the CSV DataFrame.
    """
    try:
        missing_counts = df.null_count().to_dict()
        for col, count in missing_counts.items():
            if count > 0:
                logging.info(f"CSV '{csv_path}': Column '{col}' has {count} missing values.")
    except Exception as e:
        logging.error(f"Error logging missing values for CSV '{csv_path}': {e}")

def parse_columns(df, date_column='Date'):
    """
    Parses columns to their appropriate data types.
    Specifically handles the Date column and attempts to parse numerical columns.
    Logs any parsing errors.
    """
    try:
        # Parse the Date column
        if date_column in df.columns:
            df = df.with_column(
                pl.col(date_column).str.strptime(pl.Date, fmt="%Y-%m-%d", strict=False).alias(date_column)
            )
            logging.info(f"Parsed '{date_column}' column to Date type.")
        else:
            logging.warning(f"Date column '{date_column}' not found in DataFrame.")
        
        # Attempt to parse other columns to numeric where possible
        numeric_columns = [col for col in df.columns if col != date_column]
        for col in numeric_columns:
            # Attempt to parse to Float
            df = df.with_column(
                pl.col(col).cast(pl.Float64, strict=False)
            )
        logging.info("Completed parsing columns to appropriate data types.")
        return df
    except Exception as e:
        logging.error(f"Error parsing columns: {e}")
        print(f"Error parsing columns: {e}")
        return df  # Return DataFrame even if parsing fails

def merge_data(excel_df, csv_df, date_column='Date'):
    """
    Merges CSV DataFrame into Excel DataFrame based on the date column.
    """
    try:
        # Ensure the date column is in datetime format
        if date_column in excel_df.columns:
            excel_df = excel_df.with_column(
                pl.col(date_column).cast(pl.Date, strict=False)
            )
        else:
            logging.warning(f"Date column '{date_column}' not found in Excel DataFrame.")
        
        if date_column in csv_df.columns:
            csv_df = csv_df.with_column(
                pl.col(date_column).cast(pl.Date, strict=False)
            )
        else:
            logging.warning(f"Date column '{date_column}' not found in CSV DataFrame.")
        
        # Perform a join on the date column
        merged_df = excel_df.join(csv_df, on=date_column, how='outer')
        
        logging.info("Successfully merged CSV data into Excel sheet.")
        print("Successfully merged CSV data into Excel sheet.")
        return merged_df
    except Exception as e:
        logging.error(f"Error merging data: {e}")
        print(f"Error merging data: {e}")
        return excel_df  # Return original if merge fails

def remove_oldest_rows(excel_sheets_dict, date_column='Date'):
    """
    Removes rows with the oldest date from each sheet.
    """
    for sheet_name, df in excel_sheets_dict.items():
        try:
            print(f"Processing sheet '{sheet_name}' to remove oldest rows...")
            # Check if date column exists
            if date_column not in df.columns:
                logging.warning(f"Date column '{date_column}' not found in sheet '{sheet_name}'. Skipping removal of oldest rows.")
                print(f"Date column '{date_column}' not found in sheet '{sheet_name}'. Skipping removal of oldest rows.")
                continue
            
            # Ensure the date column is of Date type
            if df[date_column].dtype != pl.Date:
                df = df.with_column(
                    pl.col(date_column).cast(pl.Date, strict=False)
                )
            
            # Find the minimum date, ignoring nulls
            min_date_series = df.select(pl.col(date_column).min()).to_series()
            min_date = min_date_series.drop_nulls().to_list()[0] if min_date_series.drop_nulls().to_list() else None
            
            if min_date:
                # Filter out rows with the minimum date
                df_filtered = df.filter(pl.col(date_column) != min_date)
                excel_sheets_dict[sheet_name] = df_filtered
                logging.info(f"Removed rows with date '{min_date}' from sheet '{sheet_name}'.")
                print(f"Removed rows with date '{min_date}' from sheet '{sheet_name}'.")
            else:
                logging.warning(f"No valid dates found in sheet '{sheet_name}'. Skipping removal of oldest rows.")
                print(f"No valid dates found in sheet '{sheet_name}'. Skipping removal of oldest rows.")
        except Exception as e:
            logging.error(f"Error removing oldest rows in sheet '{sheet_name}': {e}")
            print(f"Error removing oldest rows in sheet '{sheet_name}': {e}")

def save_final_excel(excel_sheets_dict, original_excel_path):
    """
    Saves the final Excel file with a timestamp and creates a backup of the original.
    """
    try:
        # Create backup
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{os.path.splitext(original_excel_path)[0]}_backup_{timestamp}.xlsx"
        # Copy the original file to backup instead of renaming to keep it intact
        shutil.copy2(original_excel_path, backup_path)
        logging.info(f"Backup created at '{backup_path}'")
        print(f"Backup of the original Excel file created at '{backup_path}'")
        
        # Convert Polars DataFrames to pandas DataFrames
        pandas_dict = {}
        print("Converting Polars DataFrames to pandas DataFrames...")
        for sheet, df in excel_sheets_dict.items():
            pandas_df = df.to_pandas()
            pandas_dict[sheet] = pandas_df
            logging.info(f"Converted sheet '{sheet}' to pandas DataFrame.")
            # Free memory
            del df
            gc.collect()
        
        # Save to new Excel with timestamp
        final_excel_path = f"{os.path.splitext(original_excel_path)[0]}_Final_{timestamp}.xlsx"
        print(f"Saving merged data to '{final_excel_path}'...")
        with pd.ExcelWriter(final_excel_path, engine='openpyxl') as writer:
            for sheet, df in pandas_dict.items():
                df.to_excel(writer, sheet_name=sheet, index=False)
                logging.info(f"Saved sheet '{sheet}' to final Excel file.")
                # Free memory
                del df
                gc.collect()
        logging.info(f"Final Excel file saved at '{final_excel_path}'")
        print(f"Final Excel file saved at '{final_excel_path}'")
    except Exception as e:
        logging.error(f"Error saving final Excel file: {e}")
        print(f"Error saving final Excel file: {e}")

def main():
    # Start overall timer
    overall_start = datetime.now()

    # Define current working directory
    cwd = os.getcwd()

    excel_file = os.path.join(cwd, 'NA Trend Report.xlsx')
    csv_pattern = os.path.join(cwd, "*.csv")

    # Check if Excel file exists
    if not os.path.isfile(excel_file):
        logging.error(f"Excel file '{excel_file}' not found in the current directory.")
        sys.exit(f"Excel file '{excel_file}' not found in the current directory.")

    # Read Excel sheets
    print("Reading Excel file...")
    excel_start = datetime.now()
    excel_sheets = read_excel_sheets(excel_file)
    excel_end = datetime.now()
    elapsed_excel = excel_end - excel_start
    print(f"Excel file read in {elapsed_excel}")
    logging.info(f"Excel file '{excel_file}' read in {elapsed_excel}")

    # Find all CSV files matching the pattern
    csv_files = glob.glob(csv_pattern)

    # Exclude the Excel file if it has a .csv extension
    csv_files = [f for f in csv_files if os.path.basename(f) != 'NA Trend Report.xlsx']

    if not csv_files:
        logging.error("No CSV files found in the current directory.")
        sys.exit("No CSV files found.")

    # Process each CSV file
    print("Processing CSV files...")
    for csv_file in tqdm(csv_files, desc="CSV Files"):
        sheet_name = map_csv_to_sheet(os.path.basename(csv_file))
        if not sheet_name:
            logging.warning(f"No mapping found for CSV file '{csv_file}'. Skipping.")
            print(f"No mapping found for CSV file '{csv_file}'. Skipping.")
            continue

        if sheet_name not in excel_sheets:
            logging.warning(f"Sheet '{sheet_name}' not found in Excel file. Skipping CSV '{csv_file}'.")
            print(f"Sheet '{sheet_name}' not found in Excel file. Skipping CSV '{csv_file}'.")
            continue

        csv_df = read_csv_file(csv_file)
        if csv_df is None:
            logging.error(f"Failed to read CSV file '{csv_file}'. Skipping.")
            print(f"Failed to read CSV file '{csv_file}'. Skipping.")
            continue

        # Parse columns to handle mixed data types
        csv_df = parse_columns(csv_df)

        # Merge CSV data into the corresponding Excel sheet
        merged_df = merge_data(excel_sheets[sheet_name], csv_df)
        excel_sheets[sheet_name] = merged_df
        logging.info(f"Merged CSV file '{csv_file}' into sheet '{sheet_name}'.")
        print(f"Merged CSV file '{csv_file}' into sheet '{sheet_name}'.")

        # Free memory
        del csv_df
        del merged_df
        gc.collect()

    # Remove oldest rows
    print("Removing oldest rows from each sheet...")
    remove_start = datetime.now()
    remove_oldest_rows(excel_sheets)
    remove_end = datetime.now()
    elapsed_remove = remove_end - remove_start
    print(f"Oldest rows removed in {elapsed_remove}")
    logging.info(f"Oldest rows removed in {elapsed_remove}")

    # Save final Excel file
    print("Saving final Excel file...")
    save_start = datetime.now()
    save_final_excel(excel_sheets, excel_file)
    save_end = datetime.now()
    elapsed_save = save_end - save_start
    print(f"Final Excel file saved in {elapsed_save}")
    logging.info(f"Final Excel file saved in {elapsed_save}")

    # End overall timer
    overall_end = datetime.now()
    total_time = overall_end - overall_start
    print(f"Script completed in {total_time}")
    logging.info(f"Script completed in {total_time}")

if __name__ == "__main__":
    main()
