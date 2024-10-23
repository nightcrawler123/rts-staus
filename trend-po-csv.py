import sys
import subprocess
import importlib
import os
import glob
import logging
from datetime import datetime
import shutil
from tqdm import tqdm
import gc  # For garbage collection

# ==========================
# 1. Setup and Dependencies
# ==========================

# List of required packages
required_packages = [
    'polars',
    'tqdm',
    'openpyxl',
    'pyarrow',
    'pandas'  # Added Pandas for final Excel writing
]

# Function to install missing packages
def install_packages(packages):
    for package in packages:
        try:
            importlib.import_module(package)
        except ImportError:
            print(f"Package '{package}' not found. Installing...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# Install missing packages
install_packages(required_packages)

# Now import the installed packages
import polars as pl
import openpyxl
import pandas as pd

# Setup logging at the very beginning to capture all events
logging.basicConfig(
    filename='data_processing_optimized_corrected.log',  # Updated log filename for clarity
    filemode='w',  # Overwrite log file each run
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

logging.info("Optimized Script Started.")
print("Optimized Script Started.")

# ==========================
# 2. Define CSV to Sheet Mapping
# ==========================

def map_csv_to_sheet(sheet_name):
    """
    Maps a sheet name to itself.
    Since CSVs are generated from sheet names, mapping is direct.
    """
    return sheet_name

# ==========================
# 3. Handle Duplicate Columns
# ==========================

def handle_duplicate_columns(pl_df):
    """
    Renames duplicate columns by appending suffixes to ensure uniqueness.
    For example, if 'Sales' appears twice, they become 'Sales' and 'Sales_1'.
    """
    try:
        columns = pl_df.columns
        seen = {}
        new_columns = []
        for col in columns:
            if col in seen:
                seen[col] += 1
                new_col = f"{col}_{seen[col]}"
                logging.warning(f"Duplicate column '{col}' found. Renaming to '{new_col}'.")
                print(f"Duplicate column '{col}' found. Renaming to '{new_col}'.")
            else:
                seen[col] = 0
                new_col = col
            new_columns.append(new_col)
        pl_df = pl_df.with_columns([pl.col(col).alias(new_col) for col, new_col in zip(columns, new_columns)])
        return pl_df
    except Exception as e:
        logging.error(f"Error handling duplicate columns: {e}")
        print(f"Error handling duplicate columns: {e}")
        return pl_df

# ==========================
# 4. Convert Excel Sheets to CSV
# ==========================

def convert_excel_to_csv(excel_path, temp_dir):
    """
    Converts each sheet in the Excel file to separate CSV files for faster processing.
    """
    try:
        # Ensure temp directory exists
        os.makedirs(temp_dir, exist_ok=True)
    
        # Load the Excel workbook
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        logging.info(f"Found sheets: {sheet_names}")
        print(f"Found sheets: {sheet_names}")
    
        for sheet in tqdm(sheet_names, desc="Converting Sheets to CSV"):
            ws = wb[sheet]
            # Replace any characters in sheet name that are invalid in filenames
            # Here, we assume sheet names are safe. If not, implement sanitation.
            csv_path = os.path.join(temp_dir, f"{sheet}.csv")
    
            # Read the sheet data
            data = ws.values
            try:
                cols = next(data)
            except StopIteration:
                cols = []
            with open(csv_path, 'w', encoding='utf-8') as f:
                # Write header
                f.write(','.join([str(col) if col is not None else "" for col in cols]) + '\n')
                # Write data rows
                for row in data:
                    row = [str(cell) if cell is not None else "" for cell in row]
                    f.write(','.join(row) + '\n')
    
            logging.info(f"Converted sheet '{sheet}' to CSV at '{csv_path}'")
            print(f"Converted sheet '{sheet}' to CSV at '{csv_path}'")
    
        wb.close()
    
    except Exception as e:
        logging.error(f"Error converting Excel to CSV: {e}")
        print(f"Error converting Excel to CSV: {e}")
        sys.exit(1)

# ==========================
# 5. Process CSV Files with Polars
# ==========================

def process_csv_files(temp_dir, processed_dir, original_excel_columns):
    """
    Processes each CSV file using Polars:
    - Handles duplicate columns
    - Converts date columns
    - Handles missing values
    - Removes duplicate rows
    - Aligns columns with original Excel columns
    """
    try:
        os.makedirs(processed_dir, exist_ok=True)
        csv_files = glob.glob(os.path.join(temp_dir, "*.csv"))
    
        for csv_file in tqdm(csv_files, desc="Processing CSV Files"):
            sheet_name = os.path.splitext(os.path.basename(csv_file))[0]
            target_sheet = map_csv_to_sheet(sheet_name)
    
            # Check if the target sheet exists in original_excel_columns
            if target_sheet not in original_excel_columns:
                print(f"No corresponding Excel sheet for CSV '{csv_file}'. Skipping.")
                logging.warning(f"No corresponding Excel sheet for CSV '{csv_file}'. Skipping.")
                continue
    
            # Read CSV with Polars, setting truncate_ragged_lines=True to handle extra fields
            try:
                pl_df = pl.read_csv(csv_file, truncate_ragged_lines=True)
            except Exception as e:  # Use generic exception to catch any issues
                logging.error(f"Error reading CSV '{csv_file}': {e}")
                print(f"Error reading CSV '{csv_file}': {e}")
                continue
    
            if pl_df.is_empty():
                print(f"CSV file '{csv_file}' is empty. Skipping.")
                logging.warning(f"CSV file '{csv_file}' is empty. Skipping.")
                continue
    
            # Handle duplicate columns
            pl_df = handle_duplicate_columns(pl_df)
    
            # Assume the first column is the date column
            date_column = pl_df.columns[0]
            print(f"Identified date column in CSV: '{date_column}'")
            logging.info(f"Identified date column in CSV: '{date_column}'")
    
            # Convert the date column to datetime
            pl_df = pl_df.with_columns(
                pl.col(date_column).str.strptime(pl.Date, fmt="%Y-%m-%d", strict=False).alias(date_column)
            )
    
            # Find the oldest date
            oldest_date = pl_df.select(pl.col(date_column).min()).to_series()[0]
            if oldest_date is None:
                print(f"No valid dates found in CSV '{csv_file}'. Skipping deletion.")
                logging.warning(f"No valid dates found in CSV '{csv_file}'. Skipping deletion.")
            else:
                # Filter out rows with the oldest date
                pl_df = pl_df.filter(pl.col(date_column) != oldest_date)
                logging.info(f"Deleted rows with the oldest date '{oldest_date}' from CSV '{csv_file}'.")
                print(f"Deleted rows with the oldest date '{oldest_date}' from CSV '{csv_file}'.")
    
            # Handle missing values
            for col in pl_df.columns:
                if pl_df[col].dtype in [pl.Float64, pl.Int64, pl.UInt64]:
                    mean_val = pl_df[col].mean()
                    pl_df = pl_df.with_columns(pl.col(col).fill_null(mean_val))
                elif pl_df[col].dtype == pl.Date:
                    pl_df = pl_df.with_columns(
                        pl.lit(datetime(1970, 1, 1)).cast(pl.Date).alias(col).fill_null(datetime(1970, 1, 1))
                    )
                else:
                    pl_df = pl_df.with_columns(pl.col(col).fill_null("Unknown"))
    
            # Remove duplicate rows
            initial_row_count = pl_df.height
            pl_df = pl_df.unique()
            final_row_count = pl_df.height
            duplicates_removed = initial_row_count - final_row_count
            if duplicates_removed > 0:
                print(f"Removed {duplicates_removed} duplicate rows from CSV '{csv_file}'.")
                logging.info(f"Removed {duplicates_removed} duplicate rows from CSV '{csv_file}'.")
    
            # Align columns with original Excel columns
            original_columns = original_excel_columns[target_sheet]
            csv_columns = pl_df.columns
    
            # Add missing columns with 'Unknown'
            missing_columns = set(original_columns) - set(csv_columns)
            for col in missing_columns:
                pl_df = pl_df.with_columns(pl.lit("Unknown").alias(col))
                logging.warning(f"Column '{col}' missing in CSV '{csv_file}'. Filled with 'Unknown'.")
                print(f"Column '{col}' missing in CSV '{csv_file}'. Filled with 'Unknown'.")
    
            # Reorder columns to match Excel sheet
            pl_df = pl_df.select(original_columns)
    
            # Save processed CSV
            processed_csv_path = os.path.join(processed_dir, f"{sheet_name}_processed.csv")
            pl_df.write_csv(processed_csv_path)
            logging.info(f"Processed CSV saved at '{processed_csv_path}'")
            print(f"Processed CSV saved at '{processed_csv_path}'")
    
            # Clear memory
            del pl_df
            gc.collect()
    
    except Exception as e:
        logging.error(f"Error processing CSV files with Polars: {e}")
        print(f"Error processing CSV files with Polars: {e}")
        sys.exit(1)

# ==========================
# 6. Append Processed CSVs to Original Excel Sheets
# ==========================

def append_processed_csvs_to_excel(processed_dir, final_excel_path, original_excel_columns):
    """
    Appends processed CSV data to the original Excel sheets and saves to a new Excel file.
    """
    try:
        # Load the original Excel data into Polars DataFrames
        excel_dfs = {}
        for sheet, cols in original_excel_columns.items():
            processed_csv = os.path.join(processed_dir, f"{sheet}_processed.csv")
            if os.path.exists(processed_csv):
                try:
                    pl_df = pl.read_csv(processed_csv, truncate_ragged_lines=True)
                    excel_dfs[sheet] = pl_df
                except Exception as e:  # Use generic exception to catch any issues
                    logging.error(f"Error reading processed CSV '{processed_csv}': {e}")
                    print(f"Error reading processed CSV '{processed_csv}': {e}")
                    continue
            else:
                logging.warning(f"Expected processed CSV for sheet '{sheet}' not found. Skipping.")
                print(f"Expected processed CSV for sheet '{sheet}' not found. Skipping.")
    
        # Process each processed CSV and append to the corresponding DataFrame
        processed_csv_files = glob.glob(os.path.join(processed_dir, "*_processed.csv"))
        for processed_csv in tqdm(processed_csv_files, desc="Appending to Excel Sheets"):
            sheet_name = os.path.basename(processed_csv).replace('_processed.csv', '')
            target_sheet = map_csv_to_sheet(sheet_name)
            if not target_sheet:
                print(f"No mapping found for processed CSV '{processed_csv}'. Skipping.")
                logging.warning(f"No mapping found for processed CSV '{processed_csv}'. Skipping.")
                continue
    
            if target_sheet not in excel_dfs:
                print(f"Sheet '{target_sheet}' not loaded. Skipping CSV '{processed_csv}'.")
                logging.warning(f"Sheet '{target_sheet}' not loaded. Skipping CSV '{processed_csv}'.")
                continue
    
            # Read the processed CSV
            try:
                pl_df_new = pl.read_csv(processed_csv, truncate_ragged_lines=True)
            except Exception as e:  # Use generic exception to catch any issues
                logging.error(f"Error reading processed CSV '{processed_csv}': {e}")
                print(f"Error reading processed CSV '{processed_csv}': {e}")
                continue
    
            # Append to the existing DataFrame
            excel_dfs[target_sheet] = pl.concat([excel_dfs[target_sheet], pl_df_new], how="vertical")
    
            logging.info(f"Appended data from '{processed_csv}' to sheet '{target_sheet}'.")
            print(f"Appended data from '{processed_csv}' to sheet '{target_sheet}'.")
    
            # Clear memory
            del pl_df_new
            gc.collect()
    
        # Save all DataFrames to a new Excel file
        # Convert Polars DataFrames to Pandas for writing to Excel
        with pd.ExcelWriter(final_excel_path, engine='openpyxl') as writer:
            for sheet, pl_df in excel_dfs.items():
                pd_df = pl_df.to_pandas()
                pd_df.to_excel(writer, sheet_name=sheet, index=False)
                logging.info(f"Saved sheet '{sheet}' with {pd_df.shape[0]} rows.")
                print(f"Saved sheet '{sheet}' with {pd_df.shape[0]} rows.")
    
        print(f"\nFinal Excel file saved at '{final_excel_path}'")
        logging.info(f"Final Excel file saved at '{final_excel_path}'")
    
    except Exception as e:
        logging.error(f"Error appending processed CSVs to Excel: {e}")
        print(f"Error appending processed CSVs to Excel: {e}")
        sys.exit(1)

# ==========================
# 7. Main Execution Flow
# ==========================

def main():
    try:
        # Define current working directory
        cwd = os.getcwd()
        print(f"Current Working Directory: {cwd}")
        logging.info(f"Current Working Directory: {cwd}")
    
        # Define the path to the Excel file
        excel_filename = 'NA Trend Report.xlsx'
        excel_path = os.path.join(cwd, excel_filename)
    
        # Check if the Excel file exists
        if not os.path.isfile(excel_path):
            print(f"Excel file '{excel_filename}' not found in the current directory.")
            logging.error(f"Excel file '{excel_filename}' not found in the current directory.")
            sys.exit(1)
    
        # Define temporary directories
        temp_csv_dir = os.path.join(cwd, "temp_csv")
        processed_csv_dir = os.path.join(cwd, "processed_csv")
        os.makedirs(temp_csv_dir, exist_ok=True)
        os.makedirs(processed_csv_dir, exist_ok=True)
    
        # Step 1: Convert Excel sheets to CSV for faster processing
        print("\n--- Step 1: Converting Excel Sheets to CSV ---")
        logging.info("Starting Step 1: Converting Excel Sheets to CSV")
        convert_excel_to_csv(excel_path, temp_csv_dir)
    
        # Step 2: Define original Excel sheet columns to ensure alignment
        # This assumes that the original Excel sheets have consistent columns
        # Adjust as necessary
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        original_excel_columns = {}
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            try:
                headers = next(ws.values)
                headers = [str(col).strip() if col is not None else "" for col in headers]
                original_excel_columns[sheet] = headers
            except StopIteration:
                original_excel_columns[sheet] = []
                logging.warning(f"No headers found in sheet '{sheet}'.")
                print(f"No headers found in sheet '{sheet}'.")
        wb.close()
    
        # Step 3: Process CSV files using Polars
        print("\n--- Step 2: Processing CSV Files with Polars ---")
        logging.info("Starting Step 2: Processing CSV Files with Polars")
        process_csv_files(temp_csv_dir, processed_csv_dir, original_excel_columns)
    
        # Step 4: Append processed CSVs to Excel sheets and save to a new Excel file
        print("\n--- Step 3: Appending Processed CSVs to Excel Sheets ---")
        logging.info("Starting Step 3: Appending Processed CSVs to Excel Sheets")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        final_excel_path = os.path.join(cwd, f"NA Trend Report_Final_{timestamp}.xlsx")
        append_processed_csvs_to_excel(processed_csv_dir, final_excel_path, original_excel_columns)
    
        # Optional: Clean up temporary directories
        try:
            shutil.rmtree(temp_csv_dir)
            shutil.rmtree(processed_csv_dir)
            print("\nCleaned up temporary directories.")
            logging.info("Cleaned up temporary directories.")
        except Exception as e:
            logging.warning(f"Error cleaning up temporary directories: {e}")
            print(f"Error cleaning up temporary directories: {e}")
    
        # Final message
        print("\nData processing completed successfully.")
        logging.info("Data processing completed successfully.")
    
    except Exception as e:
        logging.error(f"An unexpected error occurred in the main execution: {e}")
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
